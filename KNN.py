import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from copy import deepcopy
from numpy.linalg import norm
from matplotlib.patches import Circle


def sort_mid(x, mid, axis):
    """
    quick sort
    :param x:the n dimensions characteristic samples
    :param mid:target position
    :param axis:the comparison criteria
    """
    start, end = 0, x.shape[0] - 1
    assert 0 <= mid <= end
    while True:
        pivot = deepcopy(x[start])
        i, j = start, end
        while i < j:
            while i < j and pivot[axis] <= x[j][axis]:
                j = j - 1
            if i == j:
                break
            x[i] = x[j]
            i = i + 1
            while i < j and pivot[axis] >= x[i][axis]:
                i = i + 1
            if i == j:
                break
            x[j] = x[i]
            j = j - 1
        x[i] = pivot
        if i < mid:
            start = i + 1
        elif i > mid:
            end = i - 1
        else:
            break


class MaxHeap:
    """
    create a max_heap with k nodeplus
    """

    def __init__(self, k=0):
        self.k = k
        self.heap = []

    def insert(self, node):
        self.heap.append(node)
        pos = len(self.heap) - 1
        while pos > 0:
            parent = pos - 1 >> 1
            if node[1] <= self.heap[parent][1]:
                break
            self.heap[pos] = self.heap[parent]
            pos = parent
        self.heap[pos] = node

    def replace(self, node):
        self.heap[0] = node
        pos = 0
        son = (pos << 1) + 1
        while son <= self.k - 1:
            if son + 1 <= self.k - 1 and self.heap[son + 1][1] > self.heap[son][1]:
                son = son + 1
            if node[1] < self.heap[son][1]:
                self.heap[pos] = self.heap[son]
                pos = son
                son = (pos << 1) + 1
            else:
                break
        self.heap[pos] = node

    def push(self, node):
        if len(self.heap) < self.k:
            self.insert(node)
        else:
            self.replace(node)


class KDNode:
    def __init__(self, axis=None, data=None, label=None, parent=None, left=None, right=None, sign=0):
        self.axis = axis
        self.data = data
        self.label = label
        self.parent = parent
        self.left = left
        self.right = right
        self.sign = sign


class KDTree:
    def __init__(self, x, y):
        """
        constructor
        :param x: a matrix contains n data with k kinds of characteristics
        :param y: a array of n labels
        """
        self.root = None
        self.create(x, y)

    def create(self, x, y):
        dimension = x.shape[1]

        def create_(x_, axis, parent=None):
            """
            create kd tree
            :param x_: samples with label
            :param axis: comparison axis
            :param parent:parent node
            """
            if x_.shape[0] == 0:
                return None
            mid = x_.shape[0] >> 1
            sort_mid(x_, mid, axis)
            node = KDNode(axis, x_[mid][:-1], x_[mid][-1:], parent)
            axis = (axis+1) % dimension
            node.left = create_(x_[:mid], axis, node)
            node.right = create_(x_[mid+1:], axis, node)
            return node

        x = np.hstack((x, y))
        self.root = create_(x, 0)

    def search_knn(self, target_x, k):
        """
        find the k nearst neighbor
        """
        def arrive_leaf(target, c_knode):
            """
            :param target: the target sample
            :param c_knode: current knode
            :return: leaf knode
            """
            if c_knode is None:
                return
            while not(c_knode.left is None and c_knode.right is None):
                if target[c_knode.axis] <= c_knode.data[c_knode.axis]:
                    if c_knode.left is None:
                        c_knode = c_knode.right
                    else:
                        c_knode = c_knode.left
                else:
                    if c_knode.right is None:
                        c_knode = c_knode.left
                    else:
                        c_knode = c_knode.right
            return c_knode

        def back_off():
            c_kdnode = arrive_leaf(target_x, self.root)
            c_kdnode.sign = 1
            distance = norm(target_x - c_kdnode.data)
            kheap = MaxHeap(k)
            kheap.push((c_kdnode, distance))

            while c_kdnode.parent is not None:
                if c_kdnode.parent.sign != 1 and\
                        (len(kheap.heap) < k or norm(c_kdnode.parent.data - target_x) < kheap.heap[0][1]):
                    distance = norm(c_kdnode.parent.data - target_x)
                    kheap.push((c_kdnode.parent, distance))
                    c_kdnode.parent.sign = 1
                if len(kheap.heap) < k or \
                        abs(target_x[c_kdnode.parent.axis] - c_kdnode.parent.data[c_kdnode.parent.axis]) \
                        < kheap.heap[0][1]:
                    if c_kdnode.parent.left is not None and c_kdnode.parent.left.sign != 1:
                        c_kdnode = arrive_leaf(target_x, c_kdnode.parent.left)
                        c_kdnode.sign = 1
                        if len(kheap.heap) < k or norm(c_kdnode.data - target_x) < kheap.heap[0][1]:
                            distance = norm(c_kdnode.data - target_x)
                            kheap.push((c_kdnode, distance))
                        continue
                    if c_kdnode.parent.right is not None and c_kdnode.parent.right.sign != 1:
                        c_kdnode = arrive_leaf(target_x, c_kdnode.parent.right)
                        c_kdnode.sign = 1
                        if len(kheap.heap) < k or norm(c_kdnode.data - target_x) < kheap.heap[0][1]:
                            distance = norm(c_kdnode.data - target_x)
                            kheap.push((c_kdnode, distance))
                        continue
                c_kdnode = c_kdnode.parent
                c_kdnode.sign = 1
            return kheap
        return back_off()


def generate(workbook, worksheet, dimension):
    """
    generate n dimension data to the worksheet of the workbook.
    """
    i = 0
    while i < 100:
        x = np.random.randint(-1000, 1010, size=dimension)
        for j in range(dimension):
            worksheet.cell(i+2, j+2).value = x[j]
        if x[0] > 100 and x[1] > 100:
            worksheet.cell(i+2, dimension+2).value = 1
        elif x[0] < -100 and x[1] > 100:
            worksheet.cell(i+2, dimension+2).value = 2
        elif x[0] < -100 and x[1] < -100:
            worksheet.cell(i+2, dimension+2).value = 3
        elif x[0] > 100 and x[1] < -100:
            worksheet.cell(i+2, dimension+2).value = 4
        else:
            i = i-1
        i = i+1
    workbook.save('K_NN.xlsx')


def draw_dot_2d(worksheet, graph):
    """
    draw the data of worksheet to the graph
    """
    graph.set_xlabel("x1")
    graph.set_ylabel("x2")
    c = [[[]for i in range(2)] for j in range(4)]
    for i in range(worksheet.max_row-1):
        y = worksheet.cell(i+2, 4).value
        if y == 1:
            c[0][0].append(worksheet.cell(i+2, 2).value)
            c[0][1].append(worksheet.cell(i+2, 3).value)
        elif y == 2:
            c[1][0].append(worksheet.cell(i+2, 2).value)
            c[1][1].append(worksheet.cell(i+2, 3).value)
        elif y == 3:
            c[2][0].append(worksheet.cell(i+2, 2).value)
            c[2][1].append(worksheet.cell(i+2, 3).value)
        else:
            c[3][0].append(worksheet.cell(i+2, 2).value)
            c[3][1].append(worksheet.cell(i+2, 3).value)
    graph.scatter(c[0][0], c[0][1], c='r')
    for a, b in zip(c[0][0], c[0][1]):
        plt.text(a, b, (a, b), ha='center', va='bottom', fontsize=10)
    graph.scatter(c[1][0], c[1][1], c='g')
    for a, b in zip(c[1][0], c[1][1]):
        plt.text(a, b, (a, b), ha='center', va='bottom', fontsize=10)
    graph.scatter(c[2][0], c[2][1], c='y')
    for a, b in zip(c[2][0], c[2][1]):
        plt.text(a, b, (a, b), ha='center', va='bottom', fontsize=10)
    graph.scatter(c[3][0], c[3][1], c='blue')
    for a, b in zip(c[3][0], c[3][1]):
        plt.text(a, b, (a, b), ha='center', va='bottom', fontsize=10)


def print_tree(kdnode, graph, up, down, left, right):
    """
    draw the regions splits by the kd tree
    up, down, left, right is the limit of the sub-region
    """
    horizontal = [left, right]
    vertical = [down, up]

    if kdnode is None:
        return
    if kdnode.axis == 0:
        horizontal[0] = (kdnode.data[0])
        horizontal[1] = (kdnode.data[0])
        graph.plot(horizontal, vertical)
        print_tree(kdnode.left, graph, up, down, left, kdnode.data[0])
        print_tree(kdnode.right, graph, up, down, kdnode.data[0], right)
    elif kdnode.axis == 1:
        vertical[0] = (kdnode.data[1])
        vertical[1] = (kdnode.data[1])
        graph.plot(horizontal, vertical)
        print_tree(kdnode.left, graph, kdnode.data[1], down, left, right)
        print_tree(kdnode.right, graph, up, kdnode.data[1], left, right)


wb = openpyxl.Workbook()
wb.create_sheet(index=0, title='2-dimension')
ws = wb['2-dimension']
ws.cell(1, 2).value = 'x1'
ws.cell(1, 3).value = 'x2'
ws.cell(1, 4).value = 'y'

fig = plt.figure()
gra = fig.add_subplot(111)
generate(wb, ws, 2)

data_x = np.zeros([ws.max_row-1, 2])
data_y = np.zeros(ws.max_row-1)
for row in range(ws.max_row-1):
    for column in range(2):
        data_x[row][column] = ws.cell(row+2, column+2).value
    data_y[row] = ws.cell(row+2, 4).value

tree = KDTree(data_x, data_y.reshape(ws.max_row-1, 1))

draw_dot_2d(ws, gra)
print_tree(tree.root, gra, 1110, -1100, -1100, 1110)

# the point you wan to find
x1 = 500
x2 = 500
gra.scatter([x1], [x2], c='black')
# find k samples
ki = 10

kset = tree.search_knn(np.array([x1, x2]), ki)

distance1 = kset.heap[0][1]
leaf1 = kset.heap[0][0]
cir = Circle((x1, x2), distance1, fill=False)
gra.add_patch(cir)
gra.plot([x1, leaf1.data[0]], [x2, leaf1.data[1]], c='black')

#  show the result
for z in range(len(kset.heap)):
    print(kset.heap[z][0].data)
print(len(kset.heap))
plt.show()
